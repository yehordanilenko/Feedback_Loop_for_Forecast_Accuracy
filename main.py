from collections import defaultdict
from os.path import getctime, getmtime
from time import gmtime

import pandas as pd
from datetime import datetime
import math
from statistics import mean
from openpyxl import Workbook
from pandas import DateOffset

file_delfor = 'DELFOR 2023.12.04.csv' # Name of delfor file 04-12-2023
file_demand = 'Demand by Target SSD converter V1.0.xlsx'  # Name of demand file
NumberOfWeeksToAnalyze = 7 # Number of forecast/demand periods to analyze

data_delfor = pd.read_csv(file_delfor, sep='|', header=None, skiprows=2) # Reading data from csv
data_delfor.columns = ['Col1','Col2','Col3','Col4','Col5','PrimeItem','Col7','Col8','Col9','Col10','Qty', 'Date', 'Col13'] # Assign a name for columns

data_delfor = data_delfor[data_delfor['Col1'] != 'DEL.01'] # Deleting all rows with PrimeItem name DEL.01
data_delfor = data_delfor[data_delfor['Col1'] != 'DEL.02'] # Deleting all rows with PrimeItem name DEL.02

data_delfor = data_delfor.iloc[:, [5, 10, 11]] # Leaving only the necessary columns

text_to_add = 'CIS-'
data_delfor['PrimeItem'] = text_to_add + data_delfor['PrimeItem'].astype(str) # Concatenate CIS- to PrimeItem

# Convert columns from float to int
data_delfor['Qty'] = data_delfor['Qty'].fillna(-1).astype(int)
data_delfor['Date'] = data_delfor['Date'].fillna(-1).astype(int)

# Load the ItemSubstitutes CSV file
file_second = 'FCS_VRY_ItemSubstitutes.csv'
data_second = pd.read_csv(file_second, delimiter='|')

# Strip/trim leading and trailing spaces from the 'Item Code' column
data_second['Item Code'] = data_second['Item Code'].str.strip()


# Merge the two DataFrames to get Record No. and Priority left from delfor leftjoin on ItemCode on itemsubs
merged_data = pd.merge(data_delfor, data_second[['Item Code', 'Record No.', 'Priority']], left_on='PrimeItem', right_on='Item Code', how='left')

# Fill NaN values in 'Record No.' and 'Priority' columns
merged_data['Record No.'] = merged_data['Record No.'].fillna(0).astype(int)
merged_data['Priority'] = merged_data['Priority'].fillna(-1).astype(int)

# Create a dictionary mapping Record No. to PrimeItem with Priority 1
record_no_priority_1 = dict(data_second[data_second['Priority'] == 1].set_index('Record No.')['Item Code'])

# Apply the mapping to create a new column 'PrimeItem_Priority_1'
merged_data['PrimeItem_Priority_1'] = merged_data['Record No.'].map(record_no_priority_1)

# If 'PrimeItem_Priority_1' is NaN, use 'PrimeItem' as a fallback
merged_data['PrimeItem_Priority_1'].fillna(merged_data['PrimeItem'], inplace=True)


# Group by 'Record No.', 'Date', and 'PrimeItem_Priority_1', and aggregate
grouped_data = merged_data.groupby(['Record No.', 'Date', 'PrimeItem_Priority_1']).agg({
    'Qty': 'sum',
}).reset_index()

#check_item = 'CIS-800-110908-01'
#needed_row = grouped_data[grouped_data['PrimeItem_Priority_1'] == check_item]
#print(needed_row[['Date','Qty','PrimeItem_Priority_1']])

# Include rows with Record No. 0 in final_data without aggregation
final_data = merged_data[merged_data['Record No.'] == 0][['Record No.', 'Date', 'PrimeItem', 'Qty']].reset_index(drop=True)

# Exclude rows with Record No. 0 from further processing
grouped_data = grouped_data[grouped_data['Record No.'] != 0]

# Sort and reset index if needed
final_data = pd.concat([final_data, grouped_data], ignore_index=True).sort_values(by=['Record No.', 'Date', 'PrimeItem_Priority_1']).reset_index(drop=True)

# Drop unnecessary columns
final_data = final_data.drop(['Record No.'], axis=1)

# Fill NaN values in 'PrimeItem' with values from 'PrimeItem_Priority_1'
final_data['PrimeItem'] = final_data['PrimeItem'].fillna(final_data['PrimeItem_Priority_1'])

# Drop 'PrimeItem_Priority_1' column
final_data = final_data.drop(['PrimeItem_Priority_1'], axis=1)

# Print or use the resulting DataFrame as needed
# print(final_data)


#check_item = 'CIS-800-110908-01'
#needed_row = final_data[final_data['PrimeItem'] == check_item]
#print(needed_row[['PrimeItem','Date','Qty']])

# Print or use the resulting DataFrame as needed
# MV - Code below is to groupby prime item, to prevent duplicate item and date cobmbination as a result of primeitem translation
listx = final_data.values.tolist()

# Initialize a new list to store the filtered data
filtered_list = []

# Create a set to track seen combinations of date and itemname
seen = set()

for item in listx:
    date, itemname, qty = item

    # Construct a unique identifier for each date and itemname combination
    identifier = (date, itemname)

    # Add to filtered list if it's the first occurrence or if qty is non-zero
    if identifier not in seen or qty != 0:
        filtered_list.append(item)
        seen.add(identifier)

# Now filtered_list contains the desired data without unwanted duplicates
# If you want to calculate the total qty again, you can do so with the filtered data
total_qty = sum(item[2] for item in filtered_list)

print("====================================================================================")
print(f"Total forecast quantity: {total_qty}")
print(f"Total rows of delfor data: {len(filtered_list)}")

#all_FC_list = merged_data.values.tolist() # Convert our delfor dataframe to list
all_FC_list = filtered_list
for sublist in all_FC_list:
    sublist[0], sublist[1] = sublist[1], sublist[0]

for sublist in all_FC_list:
    sublist[1], sublist[2] = sublist[2], sublist[1]

print("Total rows of adjusted delfor data: ", len(all_FC_list))
print("====================================================================================")

for r in all_FC_list:
    date_number = r[2]
    # Convert the integer to a string and then to datetime64[ns](date type)
    date_string = str(date_number)
    date_without_time = datetime.strptime(date_string, '%Y%m%d').date()
    dd = pd.Timestamp(date_without_time)
    r[2] = dd                                   # convert int to data

# Read the Excel file into a Pandas DataFrame
data_demand = pd.read_excel(file_demand, sheet_name='TransactionHistory') #CLOSED SSD
data_demand_open = pd.read_excel(file_demand, sheet_name='PlannedTransactions') #OPEN SSD
l1 = data_demand.values.tolist()  # list of demand closed ssd file I work with it for find number of weeks
l2 = data_demand_open.values.tolist() # list of demand open ssd file I work with it to find number of weeks

sorted_list_delfor = sorted(all_FC_list, key=lambda  x: x[2]) # sorted by dates delfor

sorted_list = sorted(l1, key=lambda x: x[1]) # sorted by date demand
sorted_list2 = sorted(l2, key=lambda x: x[1])# sorted by date demand
min_date = sorted_list_delfor[0][2] if sorted_list_delfor[0][2] > sorted_list[0][1] else sorted_list[0][1] # finding minimum date  between delfor and demand

current_date = datetime.date(min_date + DateOffset(weeks=NumberOfWeeksToAnalyze-1)) # CURRENT DATE / # V1 code >> datetime.today().date() / # V2 code >> datetime(2024, 1, 7).date()

#print(l1[0])
print("Number of weeks to analyze: ", NumberOfWeeksToAnalyze)
print("Start date: ", datetime.date(min_date))
print("End date:   ", current_date) #based on first period in delfor + selected number of week
print("====================================================================================")
#Cycle for appending/grouping demand from open and closed demand. If date and item in both files, then sum qty.
list_of_demand_all = []

for i in range(len(l1)):
    for j in range(len(l2)):
        if(l1[i][0] == l2[j][0] and l1[i][1] == l2[j][1]):
            l1[i][2] += l2[j][2]

for i in range(len(l2)):
    temp_c = 0
    for j in range(len(l1)):
        if(l2[i][0] == l1[j][0] and l2[i][1] == l1[j][1]):
            temp_c += 1
    if(temp_c == 0):
        l1.append(l2[i])
list_bef_cur_date = [] # list of dates less or equal curr date
for el in l1:   # process of getting list with data where date less or equal exact(current) date
    if(el[1].date() <= current_date):
        list_bef_cur_date.append(el)

list_bef_cur_date.sort(key=lambda x: x[1])

name_sum_dict = []  # Initialize as a list

dem_cop = list_bef_cur_date.copy()

# Calculate the cumulative sum for each name
for sublist in dem_cop:
    name = sublist[0]
    value = sublist[2]

    # Check if the name is already in the list
    name_exists = False
    for entry in name_sum_dict:
        if entry[0] == name:
            entry[1] += value
            name_exists = True
            break

    # If the name is not in the list, append a new entry
    if not name_exists:
        name_sum_dict.append([name, value])

# Print the result
# print(name_sum_dict)


# Extract unique names from name_sum_dict
unique_names_in_dict = set(name for name, _ in name_sum_dict)

# Create a new list of names that are in name_sum_dict but not in list2

print("length filtered data: ", len(filtered_list))
#print([item[0] for item in filtered_list])

extra_list = [name for name in unique_names_in_dict if name not in [item[0] for item in filtered_list]]

# print(len(name_sum_dict))
# Print the result
# print(len(extra_list), extra_list)

print("Number of rows in demand file before end date: ",len(list_bef_cur_date))
print("Number of rows in demand file: ", len(l1))

sorted_list_demand_before_cur_d = sorted(list_bef_cur_date, key=lambda x: x[1]) # sorted demand list
n = sorted_list_demand_before_cur_d[len(sorted_list_demand_before_cur_d) - 1][1].date() - min_date.date() # At first n is number of days between first and last dates
n = (n/7).days + 1 # We change n to count number of weeks

print("Number of weeks between end and start date:    ", n)
print("====================================================================================")

all_rows_as_list = list_bef_cur_date # data_demand.values.tolist() # dataframe demand to list

sorted_list_delfor = sorted(all_FC_list, key=lambda x: x[2]) # Sorting delfor by date
sorted_list_demand = sorted(all_rows_as_list, key=lambda x: x[1]) # Sorting demand by date

print("Validation of rows in delfor demand list: ")
print("Rows in forecast: ", len(all_FC_list))
print("Rows in demand: ",len(all_rows_as_list))
print("====================================================================================")

spec_data = (sorted_list_demand[len(sorted_list_demand) - 1])[1] # last date in demand list
temp_data = (sorted_list_delfor[len(sorted_list_delfor) - 1])[2] # last date in delfor list

arr1 = []   # list data FC before last date in demand including this date
for i in range(len(all_FC_list)):
    if((all_FC_list[i])[2] <= spec_data ):
        arr1.append(all_FC_list[i])


print("Number of items in forecast before", current_date ,": ", len(arr1))
print("====================================================================================")

for l in arr1:    # I change rows date and qty for easier calculating in future
    l[1], l[2] = l[2], l[1]

for i in range(len(arr1)):     # I combine in delfor list qty from demand list to next process working
    for j in range(len(all_rows_as_list)):
        if((arr1[i])[0] == (all_rows_as_list[j])[0] and (arr1[i])[1] == (all_rows_as_list[j])[1]):
            arr1[i].append((all_rows_as_list[j])[2])

count = 0
setf = {''}

arr1 = sorted(arr1, key=lambda x: x[0])  # Sorting list by PrimeItem
arr2 = [item[1] for item in all_rows_as_list]
arrTe = [item[0] for item in all_rows_as_list]
finalArray = []

sumDemand = 0
countABSFCD = 0
countsq = 0
sumDemAllPeriod = 1
forecast_allPer = 0
list_temp10 = []
List_of_SKUs_with_some_D_but_0_FCST = []
List_of_SKUs_with_0_D_but_some_FCST = []
list_of_demands_and_delfors = []
#print("FIRST DATA", arr1[0])
#print(len(arr1))
#print("LAST DATA",arr1[len(arr1)-1])

#Main loop of calculation
for i in range(len(arr1)):
    if ((arr1[i])[0] not in setf):
        BIAS = count / n
        BIAS_percent = sumDemand if sumDemand == 0 else count/sumDemand # MVO 20240123: BIAS if sumDemand == 0 else count/sumDemand
        MAE = countABSFCD / n
        RMSE = math.sqrt(countsq / n)
        RMSE_percent = RMSE if sumDemAllPeriod == 0 else RMSE/(sumDemAllPeriod/n)
        SCORE = MAE + abs(BIAS)
        SCORE_percent = abs(BIAS if sumDemand == 0 else count / sumDemand) + (MAE if sumDemAllPeriod == 0 else countABSFCD / sumDemAllPeriod)
        list_of_demands_and_delfors.append([sumDemand, forecast_allPer, arr1[i][0]])
        last_for_demand = sumDemand
        last_for_FC = forecast_allPer
        finalArray.append([(arr1[i])[0], BIAS, BIAS_percent, MAE , MAE if sumDemAllPeriod == 0 else countABSFCD / sumDemAllPeriod, RMSE, RMSE_percent, SCORE, SCORE_percent])
        if(i == len(arr1)-1):
            break
        count = 0
        countsq = 0
        sumDemand = 0
        countABSFCD = 0
        sumDemAllPeriod = 0
        forecast_allPer = 0

    setf.add((arr1[i])[0])

#variables for metric calculations
    if((arr1[i])[1] in arr2):
        if(len(arr1[i]) == 4):
            count += (arr1[i])[2] - (arr1[i])[3] #diff for Bias
            countsq += ((arr1[i])[2] - (arr1[i])[3])**2
            sumDemand += (arr1[i])[3]
            countABSFCD += abs((arr1[i])[2] - (arr1[i])[3]) #diff for MAE
            sumDemAllPeriod += (arr1[i])[3] #total demand
            forecast_allPer += (arr1[i])[2] #total forecast
            if(arr1[i][0] in arrTe):
                list_temp10.append([arr1[i][0], arr1[i][1], arr1[i][2], arr1[i][3]])

        #else is case no demand
        else:
            count+= (arr1[i])[2]
            countsq+= ((arr1[i])[2])**2
            countABSFCD += abs((arr1[i])[2])
            forecast_allPer += (arr1[i])[2]
            if (arr1[i][0] in arrTe):
                list_temp10.append([arr1[i][0], arr1[i][1], arr1[i][2]])

# print("lentgh govna: ", len(list_temp10))
# for element in list_temp10:
#     print(element)

from collections import defaultdict

# Create a dictionary to store the aggregated values
aggregated_data = defaultdict(lambda: [0, 0])
for l in list_temp10:
    if(len(l) == 3):
        l.append(0)

# Iterate through the temp10 list and aggregate values by PrimeItem
for prime_item, _, fc, d in list_temp10:
    aggregated_data[prime_item][0] += fc  # Sum FC
    aggregated_data[prime_item][1] += d   # Sum D

# Convert the defaultdict to a regular dictionary
result_dict = dict(aggregated_data)

# # Print the result
# for prime_item, (total_fc, total_d) in result_dict.items():
#     print(f"PrimeItem: {prime_item}, Total FC: {total_fc}, Total D: {total_d}")
#

listik_temp = [name for name, values in result_dict.items() if all(value == 0 for value in values)]

for i in range(len(finalArray)-1):
    (finalArray[i])[1] = (finalArray[i + 1])[1]
    (finalArray[i])[2] = (finalArray[i + 1])[2]
    (finalArray[i])[3] = (finalArray[i + 1])[3]
    (finalArray[i])[4] = (finalArray[i + 1])[4]
    (finalArray[i])[5] = (finalArray[i + 1])[5]
    (finalArray[i])[6] = (finalArray[i + 1])[6]
    (finalArray[i])[7] = (finalArray[i + 1])[7]
    (finalArray[i])[8] = (finalArray[i + 1])[8]
    (list_of_demands_and_delfors[i])[0] = (list_of_demands_and_delfors[i+1])[0]
    (list_of_demands_and_delfors[i])[1] = (list_of_demands_and_delfors[i+1])[1]
    (list_of_demands_and_delfors[i])[2] = (list_of_demands_and_delfors[i + 1])[2]


(finalArray[len(finalArray)-1])[1] = count/n
(finalArray[len(finalArray)-1])[2] = BIAS_percent
(finalArray[len(finalArray)-1])[3] = countABSFCD/n
(finalArray[len(finalArray)-1])[5] = math.sqrt(countsq/n)
(finalArray[len(finalArray)-1])[6] = RMSE_percent
(finalArray[len(finalArray)-1])[7] = SCORE
(finalArray[len(finalArray)-1])[8] = SCORE_percent
(list_of_demands_and_delfors[len(list_of_demands_and_delfors)-1])[0] = last_for_demand
(list_of_demands_and_delfors[len(list_of_demands_and_delfors)-1])[1] = last_for_FC
(list_of_demands_and_delfors[len(list_of_demands_and_delfors)-1])[2] = finalArray[len(finalArray)-1][0]
print(list_of_demands_and_delfors)
print("ВОТ ВОТ ВОТ")
for i in range(len(list_of_demands_and_delfors)-1):
    (list_of_demands_and_delfors[i])[0] = (list_of_demands_and_delfors[i+1])[0]
    (list_of_demands_and_delfors[i])[1] = (list_of_demands_and_delfors[i+1])[1]
    #(list_of_demands_and_delfors[i])[2] = (list_of_demands_and_delfors[i + 1])[2]
print(list_of_demands_and_delfors)
# print("FINAL ARRAY: size is \n", len(finalArray))
# for el in finalArray:
#     print(el)


 #creating list for all items with demand > 0  and Forecast is 0
for i in range(len(list_of_demands_and_delfors)):
     if(list_of_demands_and_delfors[i][0] > 0 and list_of_demands_and_delfors[i][1] == 0):
         List_of_SKUs_with_some_D_but_0_FCST.append(list_of_demands_and_delfors[i][2])

#creating list for all items with demand = 0  and Forecast > 0
for i in range(len(list_of_demands_and_delfors)):
     if(list_of_demands_and_delfors[i][0] == 0 and list_of_demands_and_delfors[i][1] > 0):
         List_of_SKUs_with_0_D_but_some_FCST.append(list_of_demands_and_delfors[i][2])



finalArray.pop()
List_of_SKUs_with_0_D_but_some_FCST.pop()
List_of_SKUs_with_0_D_but_some_FCST.pop()
temp2 = []  # THIS FINAL LIST

for i in range(len(finalArray)):
    if((finalArray[i])[0] in arrTe):
        temp2.append(finalArray[i])
        if((temp2[len(temp2) - 1])[2] != (temp2[len(temp2) - 1])[1]):
            (temp2[len(temp2)-1])[2] = round((finalArray[i])[2], 3)
            #orignal code      (temp2[len(temp2)-1])[2] = round((finalArray[i])[2] * 100, 1)
        if((temp2[len(temp2) - 1])[4] != (temp2[len(temp2) - 1])[3]):
            (temp2[len(temp2) - 1])[4] = round((finalArray[i])[4], 3)
        if((temp2[len(temp2) - 1])[6] != (temp2[len(temp2) - 1])[5]):
            (temp2[len(temp2) - 1])[6] = round((finalArray[i])[6], 3)
        if((temp2[len(temp2) - 1])[8] != (temp2[len(temp2) - 1])[7]):
            (temp2[len(temp2) - 1])[8] = round((finalArray[i])[8], 3)

for els in temp2:
    if(els[0] in listik_temp):
        els.append("No demand and no fcst")

print("====================================================================================")

# print(temp2[(len(temp2)-1)])

#for element in temp2:
#print(element)

print("====================================================================================")
print("Metrics: ")
#print("Metrics of BIAS: ", mean([item[1] for item in temp2]))
print("Average BIAS%: ", round(mean([item[2] for item in temp2]) * 100, 1), "%")
#print("Metrics of MAE: ", mean([item[3] for item in temp2]))
print("Average MAE%: ", round(mean([item[4] for item in temp2]) * 100, 1), "%")
#print("Metrics of RMSE: ", mean([item[5] for item in temp2]))
print("Average RMSE%: ", round(mean([item[6] for item in temp2]) * 100, 1), "%")
#print("Metrics of SCORE: ", mean([item[7] for item in temp2]))
print("Average SCORE%: ", round(mean([item[8] for item in temp2]) * 100, 1), "%")
print(extra_list)


List_of_SKUs_with_demand_downside = []
List_of_SKUs_with_demand_upside = []
perfect_demand = []

#
for el in temp2:
    if(el[2] > 0.30 ):
        List_of_SKUs_with_demand_downside.append([el[0], el[2]])
    elif(el[2] < -0.30):
        List_of_SKUs_with_demand_upside.append([el[0], el[2]])
    else:
        perfect_demand.append([el[0],el[2]])

print("====================================================================================")
print(f"Number of demand downsides: {len(List_of_SKUs_with_demand_downside)}")
print(List_of_SKUs_with_demand_downside)
print(f"\nNumber of demand upsides: {len(List_of_SKUs_with_demand_upside)}")
print(List_of_SKUs_with_demand_upside)
print(f"\nNumber of good bias: {len(temp2) - len(List_of_SKUs_with_demand_downside) - len(List_of_SKUs_with_demand_upside)}")
print(perfect_demand)
print("====================================================================================")

print("List demand > 0 and FC = 0: ")
# List_of_SKUs_with_some_D_but_0_FCST.extend(extra_list)
# # Remove duplicates from original_list in place
# List_of_SKUs_with_some_D_but_0_FCST = [item for index, item in enumerate(List_of_SKUs_with_some_D_but_0_FCST) if item not in List_of_SKUs_with_some_D_but_0_FCST[:index]]

# Print the result
print(f"Length of this items: {len(List_of_SKUs_with_some_D_but_0_FCST)}")
print(List_of_SKUs_with_some_D_but_0_FCST)
print("====================================================================================")



print("List demand = 0 and FC > 0: ")
# Remove duplicates from original_list in place
List_of_SKUs_with_0_D_but_some_FCST = [item for index, item in enumerate(List_of_SKUs_with_0_D_but_some_FCST) if item not in List_of_SKUs_with_0_D_but_some_FCST[:index]]
# Print the result
print(f"Length of this items: {len(List_of_SKUs_with_0_D_but_some_FCST)}")
print(List_of_SKUs_with_0_D_but_some_FCST)
print("====================================================================================")

# Work with write in file
data = temp2
data222 = List_of_SKUs_with_demand_downside
data333 = List_of_SKUs_with_demand_upside
columns1 = ['PrimeItem','Bias','Bias%','MAE','MAE%','RMSE','RMSE%','Score','Score%', 'Comment']
# Create a new workbook and select the active worksheet
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = 'All_data'
# Write column names to Sheet 1
sheet1.append(columns1)
# Write data to the worksheet
for row in data:
    sheet1.append(row)

# Create Sheet 2
sheet2 = workbook.create_sheet(title='Problematic SKUs')
sheet2.append(['List_of_SKUs_with_demand_downside','','List_of_SKUs_with_demand_upside',''])
columns2 = ['PrimeItem (downsides)', 'BIAS% for downsides' , 'PrimeItem (upsides)', 'BIAS% for upsides']
sheet2.append(columns2)

#if(len(data222) >= len(len(data333))):
combined_list = []

for i in range(max(len(data222), len(data333))):
    item_1 = data222[i] if i < len(data222) else ["", ""]
    item_2 = data333[i] if i < len(data333) else ["", ""]
    combined_list.append(item_1 + item_2)

#print(combined_list)
# Write data to Sheet 2
for row in combined_list:
    sheet2.append(row)

sheet2.append([])

sheet3 = workbook.create_sheet(title='Items with some D but 0 FCST')
sheet3.append(['List items demand > 0 and FCST = 0'])
sheet3.append(['PrimeItem'])
for row in List_of_SKUs_with_some_D_but_0_FCST:
    sheet3.append([row])

sheet4 = workbook.create_sheet(title='Items with 0 D but some FCST')
sheet4.append(['List items demand = 0 and FCST > 0'])
sheet4.append(['PrimeItem'])
for row in List_of_SKUs_with_0_D_but_some_FCST:
    sheet4.append([row])

sheet5 = workbook.create_sheet(title='D vs FC')
sheet5.append(['PrimeItem', 'Date', 'FC', 'Demand'])

for row in list_temp10:
    sheet5.append(row)



# # Save the workbook
# filepath = 'C:/Users/vrymvolm/OneDrive - Flex/Projecten/202310 - Forecast vs Target SSD Demand/'
# filename = ' - ForecastAnalysis'
# extention = '.xlsx'
# current_date_string = str(current_date)
# CountOfWeeks = str(n)
#  workbook.save(filename= filepath + current_date_string + filename + CountOfWeeks + extention)
workbook.save(filename='Forecast Analysis test output 2.xlsx')
sheet3.append([])
